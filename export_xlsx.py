import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open(sys.argv[1], 'r', encoding='utf-8') as f:
    data = json.load(f)
rows = data['rows']
stats = data['stats']

wb = Workbook()
ws = wb.active
ws.title = "Registrations"

# Colors
PURPLE     = "6C63FF"
PURPLE_LIGHT = "EEF0FF"
DARK       = "1E1B2E"
WHITE      = "FFFFFF"
GREEN      = "27AE60"
ORANGE     = "E67E22"
BLUE       = "2980B9"
GRAY_BG    = "F5F5F5"
HEADER_BG  = "2C2A4A"
ROW_ALT    = "F9F8FF"

def side(color="CCCCCC"):
    return Side(style='thin', color=color)

def border(color="CCCCCC"):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Title block ──
ws.merge_cells("A1:V1")
title_cell = ws["A1"]
title_cell.value = "⚡ Tech Blaze 3.0 — Participant Registrations"
title_cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=PURPLE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:V2")
sub_cell = ws["A2"]
sub_cell.value = f"Exported: {data['exported_at']}   |   Total Teams: {stats['total_teams']}   |   Total Participants: {stats['total_members']}   |   Veg: {stats['veg']}   |   Non-Veg: {stats['nonveg']}"
sub_cell.font = Font(name="Arial", size=10, color="666666")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
sub_cell.fill = PatternFill("solid", fgColor=PURPLE_LIGHT)
ws.row_dimensions[2].height = 22

ws.row_dimensions[3].height = 8  # spacer

# ── Headers ──
headers = [
    "Ref ID", "Team Name", "College", "Size",
    "P1 Name", "P1 Phone", "P1 Email", "P1 Food",
    "P2 Name", "P2 Phone", "P2 Email", "P2 Food",
    "P3 Name", "P3 Phone", "P3 Email", "P3 Food",
    "P4 Name", "P4 Phone", "P4 Email", "P4 Food",
    "Medical Notes", "Registered At"
]

col_widths = [12, 18, 22, 6, 18, 14, 26, 14, 18, 14, 26, 14, 18, 14, 26, 14, 18, 14, 26, 14, 28, 20]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border("444466")
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[4].height = 28

# ── Data rows ──
for row_idx, reg in enumerate(rows, start=5):
    is_alt = (row_idx % 2 == 0)
    bg = ROW_ALT if is_alt else WHITE
    values = [
        reg.get('ref_id',''), reg.get('team',''), reg.get('college',''), reg.get('team_size',''),
        reg.get('p1',''), reg.get('p1_phone',''), reg.get('p1_email',''), reg.get('p1_food',''),
        reg.get('p2',''), reg.get('p2_phone',''), reg.get('p2_email',''), reg.get('p2_food',''),
        reg.get('p3',''), reg.get('p3_phone',''), reg.get('p3_email',''), reg.get('p3_food',''),
        reg.get('p4',''), reg.get('p4_phone',''), reg.get('p4_email',''), reg.get('p4_food',''),
        reg.get('medical',''), reg.get('created_at','')
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [7, 11, 15, 19, 21]))
        cell.border = border("DDDDDD")
        # Color Ref ID
        if col_idx == 1:
            cell.font = Font(name="Arial", size=10, bold=True, color=PURPLE)
        # Color food
        if col_idx in [8, 12, 16, 20]:
            if str(val).lower() == 'vegetarian':
                cell.font = Font(name="Arial", size=10, color=GREEN)
            elif str(val).lower() == 'non-vegetarian':
                cell.font = Font(name="Arial", size=10, color=ORANGE)
    ws.row_dimensions[row_idx].height = 20

# ── Summary Sheet ──
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 20

summary_title = ws2.cell(row=1, column=1, value="Tech Blaze 3.0 — Summary")
ws2.merge_cells("A1:B1")
summary_title.font = Font(name="Arial", size=14, bold=True, color=WHITE)
summary_title.fill = PatternFill("solid", fgColor=PURPLE)
summary_title.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

summary_data = [
    ("Total Teams Registered", stats['total_teams']),
    ("Total Participants", stats['total_members']),
    ("Vegetarian", stats['veg']),
    ("Non-Vegetarian", stats['nonveg']),
    ("Exported At", data['exported_at']),
]

for i, (label, val) in enumerate(summary_data, start=3):
    lc = ws2.cell(row=i, column=1, value=label)
    vc = ws2.cell(row=i, column=2, value=val)
    lc.font = Font(name="Arial", size=11, bold=True)
    vc.font = Font(name="Arial", size=11)
    lc.fill = vc.fill = PatternFill("solid", fgColor=PURPLE_LIGHT if i%2==0 else WHITE)
    lc.border = vc.border = border()
    lc.alignment = Alignment(vertical="center")
    vc.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[i].height = 22

out = sys.argv[1].replace('tb3_xlsx_input.json', 'techblaze_export.xlsx')
wb.save(out)
print(out)